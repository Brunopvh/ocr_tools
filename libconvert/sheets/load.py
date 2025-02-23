#!/usr/bin/env python3
#
from __future__ import annotations
import re
from typing import (List, Dict)
from pandas import (DataFrame, Series)
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from tqdm import tqdm
from abc import ABC, abstractmethod, abstractproperty
import pandas
import numpy
import math
import time
import threading

from libconvert.utils import (
    Directory,
    File,
    FilesTypes,
    InputFiles,
    FormatDate,
    FormatString,
)

#==========================================================================#
class ABCParseData(ABC):
    """
        Filtrar, editar e mainipular dados com DataFrame()
    """
    def __init__(self, df:DataFrame):
        if not isinstance(df, DataFrame):
            raise ValueError(f'{__class__.__name__} | data precisa ser do tipo DataFrame() não {type(df)}!')
        self._current_df = df.astype('str')
        self.length = len(self._current_df)
        self.current_action:str = '-'

    @property
    def data(self) -> DataFrame:
        return self._current_df
    
    @data.setter
    def data(self, new):
        if not isinstance(new, DataFrame):
            return
        self._current_df = new.astype('str')
        self.length = len(self._current_df)

    @abstractmethod
    def get_progress(self) -> float:
        pass

    @abstractmethod
    def get_text(self) -> str:
        pass

    @abstractmethod
    def header(self) -> List[str]:
        """Retorna a coluna em forma de lista."""
        pass
    
    @abstractmethod
    def exists_column(self, col:str) -> bool:
        pass
    
    @abstractmethod
    def exists_columns(self, cols:List[str]) -> bool:
        pass

    @abstractmethod
    def index_items(self, *, col:str, text:str, iqual=True) -> List[int]:
        pass
        
    @abstractmethod
    def find_elements(self, *, col:str, text:str, iqual:bool=True) -> ParseData:
        """
            Filtra texto em uma coluna, e retorna todos os elementos filtrados incluíndo
        colunas adjacentes.
        """
        pass
    
    @abstractmethod
    def find_column(self, *, col:str, text:str, iqual:bool=True) -> Series:
        pass
    
    @abstractmethod
    def select_columns(self, columns:List[str]) -> ParseData | None:
        pass
    
    @abstractmethod
    def add_column(self, *, col:str, values:List[str]) -> None:
        pass

    @abstractmethod
    def uniq_column(self, col:str) -> Series:
        pass

    @abstractmethod    
    def remove_null(self, *, col:str) -> None:
        """
            Remove valores nulos de uma coluna
        """
        pass

    @abstractmethod
    def remove_lines(self, *, col:str, text:str, iqual:bool=True) -> None:
        pass
    
    @abstractmethod
    def delet_columns(self, columns:List[str]) -> None:
        pass
    
    @abstractmethod
    def concat_columns(self, *, columns:List[str], new_col:str=None, sep_cols:str='_') -> None:
        pass

#
class ParseData(ABCParseData):
    """
        Filtrar, editar e mainipular dados com DataFrame()
    """
    def __init__(self, df):
        super().__init__(df)
        self._prog:float = 0
        self._text:str = '-'
     
    def get_progress(self):
        return self._prog
    
    def get_text(self):
        return self._text

    def header(self) -> List[str]:
        """Retorna a coluna em forma de lista."""
        return self.data.columns.tolist()
    
    def exists_column(self, col:str) -> bool:
        return col in self.header()
    
    def exists_columns(self, cols:List[str]) -> bool:
        _status = True
        for c in cols:
            if not self.exists_column(c):
                _status = False
                break
        return _status

    def index_items(self, *, col:str, text:str, iqual=True) -> List[int]:
        print(f'[OBTENDO INDICES] Coluna: {col} | Texto: {text}')
        if not self.exists_column(col):
            return []
        if iqual == False:
            return self.data[self.data[col].str.contains(text, case=False, na=False)].index.tolist()
        s: Series = self.data[col]
        return s[s == text].index.tolist()
        
    def find_elements(self, *, col:str, text:str, iqual:bool=True) -> ParseData:
        """
            Filtra texto em uma coluna, e retorna todos os elementos filtrados incluíndo
        colunas adjacentes.
        """
        print(f'[FILTRANDO TEXTO] Coluna: {col} | Texto: {text}')
        if not self.exists_column(col):
            return ParseData(DataFrame())
        #
        if iqual == True:
            df = self.data[self.data[col] == text]
        else:
            df = self.data[self.data[col].str.contains(text, case=False, na=False)]
        #
        if df.empty:
            return ParseData(DataFrame())
        #
        #list_index:List[int] = self.index_items(col=col, text=text, iqual=iqual)
        #if max(self.data.index.tolist()) < max(list_index):
        #    return ParseData(DataFrame())
        #
        #return ParseData(self.data.iloc[list_index])
        return ParseData(df)
    
    def find_column(self, *, col:str, text:str, iqual:bool=True) -> Series:
        if not self.exists_column(col):
            return Series()
        #
        if iqual == True:
            df = self.data[self.data[col] == text]
        else:
            df = self.data[self.data[col].str.contains(text, case=False, na=False)]
        #
        if df.empty:
            return Series()
        return df[col]
    
    def select_columns(self, columns:List[str]) -> ParseData | None:
        print(f'[SELECIONANDO COLUNAS] Colunas: {columns}')
        if not self.exists_columns(columns):
            return None
        return ParseData(self.data[columns])
    
    def add_column(self, *, col:str, values:List[str]) -> None:
        if self.exists_column(col):
            print(f'{__class__.__name__} [FALHA] a coluna {col} já existe no DataFrame()!')
            return
        # Preenchendo com NaN para ajustar o tamanho
        num_values:int = len(values)
        num_data:int = self.length
        if num_values == num_data:
            new_column = values
        elif num_values < num_data:
            new_column = values + [numpy.nan] * (num_data - num_values)
        elif num_values > num_data:
            new_column = values[0:num_data]
        self.data[col] = new_column
        
    def uniq_column(self, col:str) -> Series:
        return self.data[col].drop_duplicates()
        
    def remove_null(self, *, col:str) -> None:
        """
            Remove valores nulos de uma coluna
        """
        print(f'[APAGANDO LINHAS VAZIAS] Coluna: {col}')
        if not self.exists_column(col):
            return
        
        self.data = self.data.dropna(subset=[col])
        self.data = self.data[self.data[col] != "nan"]
        self.data = self.data[self.data[col] != ""]

    def remove_lines(self, *, col:str, text:str, iqual:bool=True) -> None:
        # filtered_df = df[~df['coluna'].str.contains(pattern)]
        if not self.exists_column(col):
            return
        #
        if iqual == True:
            self.data = self.data[self.data[col] != text]
        else:
            pattern = re.compile(r'{}'.format(text))
            self.data = self.data[~self.data[col].str.contains(pattern)]
    
    def delet_columns(self, columns:List[str]) -> None:
        if not self.exists_columns(columns):
            return None
        self.data = self.data.drop(columns, axis=1)
        
    def concat_columns(self, *, columns:List[str], new_col:str=None, sep_cols:str = '_'):
        """Concatena colunas no DataFrame Original"""
        # df['nova_coluna'] = df['col1'].astype(str) + '_' + df['col2'].astype(str) + '_' + df['col3'].astype(str)

        if not self.exists_columns(columns):
            print(f'As colunas não existem: {columns}')
            return None
        
        # Criar uma lista vazia para armazenar os valores concatenados
        column_concat = []

        # Iterar sobre as linhas do DataFrame
        print(f'[CONCATENANDO]: {columns}')
        iter_rows = self.data.iterrows()
        for idx, row in iter_rows:
            # Concatenar os valores das colunas selecionadas
            value_concat = sep_cols.join([str(row[col]) for col in columns])
            column_concat.append(value_concat)

        # Adicionar a nova coluna ao DataFrame
        if new_col is None:
            new_col = sep_cols.join(columns)
        self.data[new_col] = column_concat

#
class ParseDF(ABCParseData):
    """FACADE"""
    def __init__(self, df):
        super().__init__(df)
        self.parse:ABCParseData = ParseData(df)
        
    @property
    def data(self) -> DataFrame:
        return self.parse.data

    def get_progress(self) -> float:
        return self.parse.get_progress()

    def get_text(self) -> str:
        return self.parse.get_text()

    def header(self) -> List[str]:
        return self.parse.header()
    
    def exists_column(self, col:str) -> bool:
        return self.parse.exists_column(col)
    
    def exists_columns(self, cols:List[str]) -> bool:
        return self.parse.exists_columns(cols)

    def index_items(self, *, col:str, text:str, iqual=True) -> List[int]:
        return self.parse.index_items(col=col, text=text, iqual=iqual)
        
    def find_elements(self, *, col:str, text:str, iqual:bool=True) -> ParseData:
        return self.parse.find_elements(col=col, text=text, iqual=iqual)
    
    def find_column(self, *, col:str, text:str, iqual:bool=True) -> Series:
        return self.parse.find_column(col=col, text=text,iqual=iqual)
    
    def select_columns(self, columns:List[str]) -> ParseData | None:
        return self.parse.select_columns(columns)
    
    def add_column(self, *, col:str, values:List[str]) -> None:
        return self.parse.add_column(col=col, values=values)

    def uniq_column(self, col:str) -> Series:
        return self.parse.uniq_column(col)
    
    def remove_null(self, *, col:str) -> None:
        return self.parse.remove_null(col=col)

    def remove_lines(self, *, col:str, text:str, iqual:bool=True) -> None:
        return self.parse.remove_lines(col=col, text=text, iqual=iqual)
    
    def delet_columns(self, columns:List[str]) -> None:
        return self.parse.delet_columns(columns)
    
    def concat_columns(self, *, columns, new_col = None, sep_cols = '_'):
        return self.parse.concat_columns(columns=columns, new_col=new_col, sep_cols=sep_cols)


#==========================================================================#

class LoadSheet(ABC):
    def __init__(self, file:File, *, separator:str='\t', load_now:bool=False):
        super().__init__()
        if not file.is_sheet():
            raise ValueError(f'{__class__.__name__}\n[FALHA] planilha inválida!')
        self.file = file
        self.load_now:bool=load_now
        # Mostrar o progresso e textos durante a leitura dos dados.
        self.current_progress:float = 0
        self.current_text:str = '-'
        # DataFrame e cabeçalho
        self._current_df: DataFrame = None
        self._origin_header:List[str] = None
        # Controle da execução da Thread
        self.running = False
        self._main_thread:threading.Thread = None
        self._current_event:threading.Event = None
        self.current_progress = 0
        self.current_text = '-'
        
    @property
    def data(self) -> DataFrame:
        return self._current_df
    
    @data.setter
    def data(self, new):
        if not isinstance(new, DataFrame):
            return
        self._current_df = new
        
    @abstractmethod
    def start_values(self, *, inthread:bool=False):
        pass

    @abstractmethod
    def get_num_rows(self) -> int:
        pass

    @abstractmethod
    def header(self) -> List[str]:
        pass

    @abstractmethod
    def get_progress(self) -> float:
        pass

    @abstractmethod
    def get_text(self) -> str:
        pass

    @abstractmethod
    def get_data(self) -> DataFrame:
        pass

    @abstractmethod
    def stop_load(self):
        pass

class LoadCsv(LoadSheet):
    def __init__(self, file, *, separator = '\t', load_now = False):
        super().__init__(file, separator=separator, load_now=load_now)
        self.separator:str = separator
        self._num_rows:int = None
        self._current_event: threading.Event = None
        if self.load_now == True:
            self.start_values(inthread=True)

    def get_progress(self) -> str:
        if self.current_progress is None:
            return '0%'
        return f'{self.current_progress:1f}%'

    def get_num_rows(self) -> int:
        if self._num_rows is None:
            with open(self.file.absolute(), 'r', encoding='utf-8') as read_file:
                self._num_rows = sum(1 for _ in read_file)
        return self._num_rows
    
    def header(self):
        if self._origin_header is None:
            self._origin_header = self.get_data().columns.tolist()
        return self._origin_header

    def _run_start(self):
        self.running = True  
        self.current_text = f'Lendo o arquivo: {self.file.basename()}'
        print(self.current_text)  
        num_chunk = 1000
        elements:List[object] = []
        try:
            # Total de iterações no arquivo durante a leitura.
            total_iter = math.ceil(self.get_num_rows() / num_chunk)
            counter:int = 0
            #
            with pandas.read_csv(
                                self.file.absolute(), 
                                chunksize=num_chunk, 
                                encoding='utf-8', 
                                sep=self.separator
                            ) as _reader_csv:
                for chunk in _reader_csv:
                    elements.append(chunk)
                    self.current_progress = (counter/total_iter) * 100
                    self.current_text = f'Lendo: {self.file.basename()} | {self.file.size()}Kb'
                    print(f'{self.current_progress:.2f} {self.current_text}', end='\r')
                    counter += 1
        except Exception as e:
           print(e)
           self.current_text = e
        else:
            pass
            
        print()  
        if len(elements) < 1:
            self._current_df = DataFrame()
            self._origin_header = []
        else:   
            self._current_df = pandas.concat(elements, ignore_index=True)
            self._origin_header = self._current_df.columns.tolist()
        self.current_progress = 100
        self.current_text = 'OK'
        self.running = False

    def start_values(self, *, inthread = False):
        if self.running == True:
            count:int=0
            while True:
                print(f'{__class__.__name__} AGUARDANDO... {count}', end='\r')
                time.sleep(1)
                if self.running == False:
                    print()
                    break
                count += 1
        #
        if self._current_df is not None:
            self.running = False
            return
        if inthread == True:
            th = threading.Thread(target=self._run_start)
            th.start()
        else:
            self._run_start()

    def stop_load(self):
        if self._current_event is None:
            return
        self._current_event.set()
                
    def get_data(self):
        if self._current_df is None:
            self.start_values()
        return self._current_df
    
    def get_text(self):
        return self.current_text


class LoadExcel(LoadSheet):
    def __init__(self, file, *, separator = '\t', load_now = False):
        super().__init__(file, separator=separator, load_now=load_now)
        self._num_rows:int = None
        self.__read_only_wsheet:ReadOnlyWorksheet = None
        self.__origin_header:List[str] = None
        if self.load_now == True:
            self.start_values(inthread=True)
        
    def stop_load(self):
        if self._current_event is None:
            return
        self._current_event.set()
        
    def header(self) -> List[str]:
        """Cabeçalho original da planilha"""
        if self.__origin_header is None:
            self.__origin_header = next(self.get_read_only_ws().iter_rows(values_only=True))
        return self.__origin_header
    
    def get_num_rows(self):
        if self._num_rows is None:
            self.current_text = f'Calculando número de linhas: {self.file.basename()}'
            self._num_rows = self.get_read_only_ws().max_row
        return self._num_rows
    
    def _get_read_only_wb(self) -> Workbook | None:
        self.current_text = f'Lendo:\n{self.file.basename()}'
        try:
            return load_workbook(self.file.absolute(), read_only=True)
        except Exception as e:
            print()
            print(f'{__class__.__name__}\n{e}\n')
            self.current_text = f'Erro ao tentar ler o arquivo: {self.file.basename()}'
            return None
        
    def get_read_only_ws(self) -> ReadOnlyWorksheet | None:
        """Objeto para obter algumas propriedades da planilha."""
        if self.__read_only_wsheet is None:
            _wb:Workbook = self._get_read_only_wb()
            if _wb is None:
                return None
            self.__read_only_wsheet = _wb.active
        self.current_text = '-'
        return self.__read_only_wsheet
    
    def _run_start(self):
        self.running = True
        self.current_text = f'Analizando a planilha: {self.file.basename()}'
        #
        self._current_df = None
        currentData:List[tuple] = []
        if self.get_read_only_ws() is None:
            self.running = False
            return
        #
        _rows = self.get_read_only_ws().iter_rows(values_only=True)
        for num, row in enumerate(_rows):
            if num == 0:
                # Ignorar a primeira linha (cabeçalho).
                continue
            currentData.append(row)
            self.current_progress = (num/self.get_num_rows()) * 100
            self.current_text = f'Lendo planilha: {self.file.basename()}'
            print(f'{self.current_progress:.1f} | {self.current_text}', end='\r')
        #
        if len(currentData) < 1:
            self.running = False
            self._current_df = DataFrame()
        else:
            self._current_df = DataFrame(currentData)
        #
        self._current_df.columns = self.header()
        self.current_progress = 100
        self.running = False
        
    def start_values(self, *, inthread = False):
        """
            Carregar alguns valores da planilha e atualizar o 
        progresso durante a operação.
        """
        if self.running == True:
            count:int=0
            while True:
                #print(f'{__class__.__name__} AGUARDANDO... {count}')
                time.sleep(1)
                if self.running == False:
                    print()
                    break
                count += 1
        
        #
        if self._current_df is not None:
            self.running = False
            return
        if inthread == True:
            th = threading.Thread(target=self._run_start)
            th.start()
        else:
            self._run_start()

    def get_data(self) -> DataFrame:
        """
            Retorna o DataFrame da planilha.
        """
        if self._current_df is None:
            self.start_values()
        return self._current_df

    def get_text(self):
        return self.current_text
    
    def get_progress(self) -> float:
        return self.current_progress

#==========================================================================#

class GetLoadSheetFacade(object):
    def __init__(self, file:File, *, separator = '\t', load_now = False):

        if not file.is_sheet():
            raise ValueError(f'{__class__.__name__}\n Planilha inválida!')
        self.file = file
        self.separator = separator
        self.load_now = load_now
          
    def get(self) -> LoadSheet:  
        if self.file.extension() in FilesTypes.EXCEL.value:
            return LoadExcel(self.file, separator=self.separator, load_now=self.load_now)
        elif self.file.extension() in FilesTypes.CSV.value:
            return LoadCsv(self.file, separator=self.separator, load_now=self.load_now)
        else:
            raise ValueError(f'{__class__.__name__}\nArquivo inválido!') 


class SheetInputStream(LoadSheet):
    def __init__(self, file, *, separator = '\t', load_now = False):
        super().__init__(file, separator=separator, load_now=load_now)
       
        if self.file.extension() in FilesTypes.EXCEL.value:
            self._load_sheet:LoadSheet = LoadExcel(self.file, separator=separator, load_now=self.load_now)
        elif self.file.extension() in FilesTypes.CSV.value:
            self._load_sheet:LoadSheet = LoadCsv(self.file, separator=separator, load_now=self.load_now)
        else:
            raise ValueError(f'{__class__.__name__} | Arquivo inválido! | {type(file)}') 
        
    def stop_load(self):
        self._load_sheet.stop_load() 

    def is_running(self) -> bool:
        return self._load_sheet.running   
    
    def start_values(self, *, inthread = False):
        self._load_sheet.start_values(inthread=inthread)

    def get_progress(self):
        return self._load_sheet.get_progress()
    
    def get_text(self):
        return self._load_sheet.get_text()

    def header(self):
        return self._load_sheet.header()

    def get_num_rows(self):
        return self._load_sheet.get_num_rows()
    
    def get_data(self) -> DataFrame:
        return self._load_sheet.get_data()
    
#==========================================================================#

    